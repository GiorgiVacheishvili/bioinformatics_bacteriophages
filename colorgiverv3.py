import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the data files
genes = input("Write the name of the genes file with extension: ")
genes_df = pd.read_excel(genes)
phrogs_df = pd.read_excel("PHROG table.xlsx")

# Dictionary to store counts of each category
category_counts = {
    'dna_packaging_and_head': 0,
    'viral_structural_proteins': 0,
    'dna_rna_metabolism': 0,
    'lysis': 0,
    'aux': 0,
    'transcription_regulation': 0,
    'head_to_tail': 0,
    'tail': 0,
    'membrane_protein': 0,
    'hypotheticals': 0
}

# Define the function to determine the color based on the logic
def assign_color(row):
    protein = row["Protein"]
    matching_row = phrogs_df[phrogs_df['Annotation'] == protein]

    # Check if a matching row was found
    if not matching_row.empty:
        category = matching_row.iloc[0]['Category'].lower()
    else:
        category = ""

    if pd.notna(protein):
        protein_lower = protein.lower()

        if 'hypothetical protein' in protein_lower:
            category_counts['hypotheticals'] += 1
            return 'C3C49E'
        elif 'membrane protein' in protein_lower:
            category_counts['membrane_protein'] += 1
            return 'CBC0D3'
        elif 'aux' in category:
            category_counts['aux'] += 1
            return '3D405B'
        elif 'vir' in protein_lower:
            category_counts['viral_structural_proteins'] += 1
            return "63CCCA"
        elif 'packag' in category:
            category_counts['dna_packaging_and_head'] += 1
            return '4F7CAC'
        elif 'head-tail' in protein_lower:
            category_counts['head_to_tail'] += 1
            return '06A77D'
        elif 'tail' in protein_lower or 'tail' in category:
            category_counts['tail'] += 1
            return '6BAA75'
        elif 'metabolism' in category:
            category_counts['dna_rna_metabolism'] += 1
            return 'F4743B'
        elif 'lysis' in category:
            category_counts['lysis'] += 1
            return '883955'
        elif 'transcription regulation' in category:
            category_counts['transcription_regulation'] += 1
            return 'FCD757'
        elif "connector" in category:
            category_counts['head_to_tail'] += 1
            return '06A77D'
        elif "other" in category:
            category_counts['aux'] += 1
            return '3D405B'
    else:
        return ""

# Apply the function to get the color codes
genes_df['Color'] = genes_df.apply(assign_color, axis=1)

# Load the Excel file using openpyxl
wb = load_workbook(genes)
ws = wb.active

# Apply the color to each row in the Excel sheet
for index, row in genes_df.iterrows():
    color_code = row['Color']
    if color_code:  # Only apply fill if color_code is not empty
        fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
        for col_num in range(1, len(genes_df.columns)):  # Apply color to all columns except 'Color'
            ws.cell(row=index + 2, column=col_num).fill = fill

# Save the result to a new Excel file
wb.save('ColoredGenes.xlsx')

# Print out the counts of each category
for category, count in category_counts.items():
    print(f"{category.replace('_', ' ').capitalize()}: {count}")
